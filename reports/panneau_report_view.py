"""
Rapport d'état des panneaux — vue, contexte et export PDF/Excel.
"""

from datetime import date, datetime
import io

import pandas as pd
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.views import View
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from weasyprint import HTML

from accounts.decorators import ClientStaffRequiredMixin
from campaigns.models import Campagne, Client, ReservationLigne,STATUT_EN_ATTENTE, STATUT_CONFIRMEE
from inventory.models import *
from django.utils import timezone
from django.db.models import Q


# ══════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════

OCCUPATION_CHOICES = {
    'occupe'          : 'Occupé (au moins une face avec campagne)',
    'libre'           : 'Libre (aucune face occupée ni réservée)',
    'reserve'         : 'Réservé (au moins une face réservée)',
    'total_occupe'    : 'Totalement occupé (toutes faces avec campagne)',
    'total_reserve'   : 'Totalement réservé (toutes faces réservées)',
    'non_reserve'     : 'Non réservé (aucune réservation)',
    'occupe_ou_reserve': 'Occupé ou réservé',
}


# ══════════════════════════════════════════════════════════════════
# UTILITAIRES
# ══════════════════════════════════════════════════════════════════



def _occupation_label(faces_data):
    nb_total   = len(faces_data)
    nb_panne   = sum(1 for f in faces_data if f['statut'] == 'panne')
    nb_occupe  = sum(1 for f in faces_data if f['statut'] == 'occupe')
    nb_reserve = sum(1 for f in faces_data if f['statut'] == 'reserve')

    if nb_total == 0:
        return 'libre'
    if nb_panne == nb_total:
        return 'panne'                          # toutes les faces en panne
    if nb_occupe == nb_total:
        return 'total_occupe'
    if nb_reserve == nb_total and nb_occupe == 0:
        return 'total_reserve'
    if nb_occupe > 0:
        return 'occupe'
    if nb_reserve > 0:
        return 'reserve'
    return 'libre'


def _match_filter(occupation_label, filter_occupation):
    """Vérifie si le label d'occupation du support correspond au filtre choisi."""
    if not filter_occupation:
        return True

    mapping = {
        'occupe'          : lambda o: o in ('occupe', 'total_occupe'),
        'libre'           : lambda o: o == 'libre',
        'reserve'         : lambda o: o in ('reserve', 'total_reserve'),
        'total_occupe'    : lambda o: o == 'total_occupe',
        'total_reserve'   : lambda o: o == 'total_reserve',
        'non_reserve'     : lambda o: o in ('libre', 'occupe'),
        'occupe_ou_reserve': lambda o: o != 'libre',
    }
    fn = mapping.get(filter_occupation)
    return fn(occupation_label) if fn else True


# ══════════════════════════════════════════════════════════════════
# CONSTRUCTION DU CONTEXTE
# ══════════════════════════════════════════════════════════════════

def _build_context_panneaux(filters: dict) -> dict:
    today = date.today()
    now   = timezone.now()

    # ── Base queryset ────────────────────────────────────────────
    qs = (
        Support.objects
        .filter(type_support="panneau")
        .prefetch_related("faces__lignes_campagne__campagne__client")
        .order_by("ville", "quartier", "nom")
    )

    # ── Filtres simples ──────────────────────────────────────────
    ville     = filters.get("ville",     "").strip()
    quartier  = filters.get("quartier",  "").strip()
    statut    = filters.get("statut",    "").strip()
    q         = filters.get("q",         "").strip()
    client_pk = filters.get("client_pk", "")

    if ville:
        qs = qs.filter(ville__iexact=ville)
    if quartier:
        qs = qs.filter(quartier__iexact=quartier)
    if statut:
        qs = qs.filter(etat=statut)
    if q:
        qs = (
            qs.filter(nom__icontains=q)
            | qs.filter(code__icontains=q)
            | qs.filter(adresse__icontains=q)
        )
    # Le filtre client est plus complexe car il nécessite de vérifier les campagnes actives sur les faces , meme celles reservees, pour éviter de montrer un panneau comme libre alors qu'il est réservé pour un client donné
    if client_pk:
        qs = qs.filter(
            Q(lignes_campagne__campagne__client_id=client_pk) |  # faces avec campagne active
            Q(faces__lignes_reservation__reservation__client_id=client_pk)       # faces réservées directement
        ).distinct()

    # ── Pré-chargement des faces réservées (1 seule requête SQL) ─
    faces_reservees_ids = set(
        ReservationLigne.objects
        .filter(
            reservation__date_fin__gte=now,
            reservation__statut__in=[STATUT_EN_ATTENTE, STATUT_CONFIRMEE],
        )
        .values_list('face_id', flat=True)
    )

    # ── Construction des lignes ──────────────────────────────────
    filter_occupation = filters.get("occupation", "").strip()

    panneaux       = []
    total_faces    = 0
    total_occupees = 0
    total_reservees = 0
    total_panne = 0
    total_libres = 0

    for support in qs:
        faces_data = []

        for face in support.faces.all():
            campagnes_list = list(
                Campagne.objects
                .filter(lignes__face=face, date_fin__gte=today)
                .select_related("client")
                .distinct()
                .order_by("date_debut")
            )
            statut_face = face.get_statut()

            faces_data.append({
                "face"     : face,
                "label"    : face.label,
                "statut"   : statut_face,           # 'occupe' | 'reserve' | 'libre'
                "occupee"  : statut_face == 'occupe',
                "reservee" : statut_face == 'reserve',
                "en_panne" : statut_face == 'panne',
                "campagnes": campagnes_list,
            })

        nb_faces_total = len(faces_data)
        nb_en_panne    = sum(1 for f in faces_data if f['statut'] == 'panne')
        nb_occupees    = sum(1 for f in faces_data if f['statut'] == 'occupe')
        nb_reservees   = sum(1 for f in faces_data if f['statut'] == 'reserve')
        nb_libres      = sum(1 for f in faces_data if f['statut'] == 'libre')
        occupation     = _occupation_label(faces_data)

        # Appliquer le filtre occupation
        if not _match_filter(occupation, filter_occupation):
            continue

        total_faces     += nb_faces_total
        total_occupees  += nb_occupees
        total_reservees += nb_reservees
        total_panne += nb_en_panne
        total_libres += nb_libres

        panneaux.append({
            "support"     : support,
            "code"        : support.code,
            "nom"         : support.nom,
            "ville"       : support.ville,
            "quartier"    : support.quartier,
            "adresse"     : support.adresse,
            "statut"      : support.get_etat_display() if hasattr(support, "get_etat_display") else support.etat,
            "etat_raw"    : support.etat,
            "nb_faces"    : nb_faces_total,
            "nb_occupees" : nb_occupees,
            "nb_reservees": nb_reservees,
            "nb_en_panne" : nb_en_panne,
            "nb_libres"   : nb_libres,
            "occupation"  : occupation,
            "faces"       : faces_data,
        })

    # ── Données pour les <select> de filtres ────────────────────
    all_supports = Support.objects.filter(type_support="panneau")

    villes    = sorted({v.strip() for v in all_supports.values_list("ville",    flat=True) if v and v.strip()})
    quartiers = sorted({q.strip() for q in all_supports.values_list("quartier", flat=True) if q and q.strip()})
    statuts   = sorted({s.strip() for s in all_supports.values_list("etat",     flat=True) if s and s.strip()})
    clients   = Client.objects.order_by("nom")

    return {
        "panneaux"          : panneaux,
        "today"             : today,
        "filters"           : filters,
        "villes"            : villes,
        "quartiers"         : quartiers,
        "statuts"           : statuts,
        "clients"           : clients,
        "occupation_choices": OCCUPATION_CHOICES,
        "total_panneaux"    : len(panneaux),
        "total_faces"       : total_faces,
        "total_occupees"    : total_occupees,
        "total_reservees"   : total_reservees,
        "total_panne" : total_panne,
        "total_libres"      : total_libres,
    }


def _filters_from_request(request) -> dict:
    return {
        "ville"     : request.GET.get("ville",     ""),
        "quartier"  : request.GET.get("quartier",  ""),
        "statut"    : request.GET.get("statut",    ""),
        "occupation": request.GET.get("occupation",""),
        "client_pk" : request.GET.get("client_pk", ""),
        "q"         : request.GET.get("q",         ""),
    }

# ══════════════════════════════════════════════════════════════════
# VUES
# ══════════════════════════════════════════════════════════════════

class PanneauxReportView(ClientStaffRequiredMixin, View):
    """Aperçu HTML du rapport — inclut les filtres."""

    def get(self, request):
        filters = _filters_from_request(request)
        context = _build_context_panneaux(filters)
        return render(request, "reports/apercu_panneaux.html", context)


class ExportPanneauxPdfView(ClientStaffRequiredMixin, View):
    """Télécharge le rapport filtré en PDF."""

    def get(self, request):
        filters = _filters_from_request(request)
        context = _build_context_panneaux(filters)

        html_string = render_to_string(
            "reports/panneaux_report_pdf.html",
            context,
            request=request,
        )
        pdf = HTML(
            string=html_string,
            base_url=request.build_absolute_uri(),
        ).write_pdf()

        # Nom de fichier reflétant les filtres actifs
        parts = ["etat_panneaux"]
        if filters.get("ville"):      parts.append(filters["ville"].replace(" ", "_"))
        if filters.get("quartier"):   parts.append(filters["quartier"].replace(" ", "_"))
        if filters.get("occupation"): parts.append(filters["occupation"])
        parts.append(f"{datetime.now():%Y%m%d}")
        # ajouter HMS
        parts.append(f"{datetime.now():%H%M%S}")
        filename = "_".join(parts) + ".pdf"

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response






import io
import re
from datetime import date, datetime

import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from django.views import View
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', '_', name)
    return cleaned[:31]


def _thin_border(color="E2E8F0"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


class ExportPanneauxExcelView(ClientStaffRequiredMixin, View):

    def get(self, request):
        filters = _filters_from_request(request)
        context = _build_context_panneaux(filters)
        now     = timezone.now()

        # ── Nombre max de faces ──
        max_faces = max(
            (len(panneau["faces"]) for panneau in context["panneaux"]),
            default=1
        )

        # ── Colonnes sheet principal ──
        base_cols = ["Code", "Nom", "Ville", "Quartier", "Adresse", "Statut"]
        face_cols = [f"Face {chr(65 + i)}" for i in range(max_faces)]
        all_cols  = base_cols + face_cols

        # ── Lignes + collecte données clients ──
        rows         = []
        clients_data = {}

        for panneau in context["panneaux"]:
            row = {
                "Code"    : panneau["code"],
                "Nom"     : panneau["nom"],
                "Ville"   : panneau["ville"],
                "Quartier": panneau["quartier"],
                "Adresse" : panneau["adresse"],
                "Statut"  : panneau["statut"],
            }

            for i, face_data in enumerate(panneau["faces"]):
                lettre = chr(65 + i)
                statut = face_data["statut"]

                if statut == "occupe" and face_data["campagnes"]:
                    camp       = face_data["campagnes"][0]
                    client_nom = camp.client.nom if camp.client else "Inconnu"
                    contenu    = f"● Occupée\n{client_nom} — {camp.nom}"
                    if client_nom not in clients_data:
                        clients_data[client_nom] = []
                    clients_data[client_nom].append({
                        "Code"    : panneau["code"],
                        "Nom"     : panneau["nom"],
                        "Ville"   : panneau["ville"],
                        "Quartier": panneau["quartier"],
                        "Adresse" : panneau["adresse"],
                        "Face"    : lettre,
                        "Statut"  : "Occupée",
                        "Campagne": camp.nom,
                        "Début"   : camp.date_debut.strftime("%d/%m/%Y"),
                        "Fin"     : camp.date_fin.strftime("%d/%m/%Y"),
                    })

                elif statut == "reserve":
                    resa = (
                        face_data["face"].lignes_client
                        .filter(date_fin__gte=now)
                        .select_related("client")
                        .first()
                    )
                    client_nom = resa.client.nom if resa and resa.client else "Inconnu"
                    contenu    = f"◈ Réservée\n{client_nom}" if resa else "◈ Réservée"
                    if resa and resa.client:
                        if client_nom not in clients_data:
                            clients_data[client_nom] = []
                        clients_data[client_nom].append({
                            "Code"    : panneau["code"],
                            "Nom"     : panneau["nom"],
                            "Ville"   : panneau["ville"],
                            "Quartier": panneau["quartier"],
                            "Adresse" : panneau["adresse"],
                            "Face"    : lettre,
                            "Statut"  : "Réservée",
                            "Campagne": "— (Réservation)",
                            "Début"   : resa.date_debut.strftime("%d/%m/%Y"),
                            "Fin"     : resa.date_fin.strftime("%d/%m/%Y"),
                        })

                else:
                    contenu = "__"

                row[f"Face {lettre}"] = contenu

            rows.append(row)

        df = pd.DataFrame(rows, columns=all_cols)

        # ── Styles Design 2 — Moderne contrasté ──
        # En-têtes
        header_fill = PatternFill("solid", start_color="0F172A", end_color="0F172A")
        header_font = Font(bold=True, color="E2E8F0", name="Arial", size=10)

        # Titre
        titre_font  = Font(bold=True, size=13, name="Arial", color="0F172A")

        # Faces
        fill_occupe   = PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")
        font_occupe   = Font(bold=True, color="166534", name="Arial", size=10)

        fill_reserve  = PatternFill("solid", start_color="EDE9FE", end_color="EDE9FE")
        font_reserve  = Font(bold=True, color="5B21B6", name="Arial", size=10)

        fill_libre    = PatternFill("solid", start_color="F1F5F9", end_color="F1F5F9")
        font_libre    = Font(color="64748B", name="Arial", size=10)

        # En panne : toute la ligne
        fill_panne_row  = PatternFill("solid", start_color="FFF7ED", end_color="FFF7ED")
        fill_panne_face = PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")
        font_panne_row  = Font(color="9A3412", name="Arial", size=10)
        font_panne_face = Font(bold=True, color="991B1B", name="Arial", size=10)

        # Lignes normales
        fill_normal = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        font_normal = Font(color="0F172A", name="Arial", size=10)

        border_normal = _thin_border("E2E8F0")
        border_panne  = _thin_border("FED7AA")

        # ── Export Excel ──
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

            # ════════════════════════════════
            # SHEET PRINCIPAL — État Panneaux
            # ════════════════════════════════
            df.to_excel(writer, index=False, sheet_name="État Panneaux", startrow=2)
            ws = writer.sheets["État Panneaux"]

            # Titre ligne 1
            titre_parts = [f"État des panneaux — généré le {date.today():%d/%m/%Y}"]
            if filters.get("ville"):      titre_parts.append(f"Ville : {filters['ville']}")
            if filters.get("quartier"):   titre_parts.append(f"Quartier : {filters['quartier']}")
            if filters.get("occupation"): titre_parts.append(f"Occupation : {filters['occupation']}")
            ws["A1"] = "  |  ".join(titre_parts)
            ws["A1"].font = titre_font

            # En-têtes ligne 3
            for cell in ws[3]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = _thin_border("1E293B")

            ws.row_dimensions[3].height = 22

            # Index colonnes
            face_col_indices = [i + 1 for i, col in enumerate(all_cols) if col.startswith("Face ")]
            statut_col_index = all_cols.index("Statut") + 1

            # ── Coloration par ligne ──
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
                statut_val = row[statut_col_index - 1].value
                en_panne   = statut_val and statut_val != "Bon état"

                for cell in row:
                    cell.font   = font_panne_row if en_panne else font_normal
                    cell.fill   = fill_panne_row if en_panne else fill_normal
                    cell.border = border_panne   if en_panne else border_normal
                    cell.alignment = Alignment(vertical="center", wrap_text=False)

                # Surcharge cellules faces
                for idx in face_col_indices:
                    cell = row[idx - 1]
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

                    if en_panne:
                        cell.fill = fill_panne_face
                        cell.font = font_panne_face
                        if cell.value and "Libre" in str(cell.value):
                            cell.value = "⚠ En panne"
                        elif cell.value:
                            cell.value = "⚠ En panne"
                    else:
                        val = str(cell.value) if cell.value else ""
                        if "Occupée" in val:
                            cell.fill = fill_occupe
                            cell.font = font_occupe
                        elif "Réservée" in val:
                            cell.fill = fill_reserve
                            cell.font = font_reserve
                        else:
                            cell.fill = fill_libre
                            cell.font = font_libre

                ws.row_dimensions[row[0].row].height = 30

            # Largeurs colonnes
            col_widths = {
                "Code": 10, "Nom": 50, "Ville": 15,
                "Quartier": 15, "Adresse": 35, "Statut": 14,
            }
            for i, col_name in enumerate(all_cols, start=1):
                ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 28)

            ws.freeze_panes = "G4"

            # ════════════════════════════════
            # SHEETS PAR CLIENT
            # ════════════════════════════════
            for client_nom, lignes in sorted(clients_data.items()):
                sheet_name = _sanitize_sheet_name(client_nom)

                panneaux_client = {}
                for ligne in lignes:
                    code = ligne["Code"]
                    if code not in panneaux_client:
                        panneaux_client[code] = {
                            "Code"    : ligne["Code"],
                            "Nom"     : ligne["Nom"],
                            "Ville"   : ligne["Ville"],
                            "Quartier": ligne["Quartier"],
                            "Adresse" : ligne["Adresse"],
                        }
                    lettre = ligne["Face"]
                    if ligne["Statut"] == "Occupée":
                        panneaux_client[code][f"Face {lettre}"] = (
                            f"● Occupée\n{ligne['Campagne']}\n{ligne['Début']} → {ligne['Fin']}"
                        )
                    else:
                        panneaux_client[code][f"Face {lettre}"] = (
                            f"◈ Réservée\n{ligne['Début']} → {ligne['Fin']}"
                        )

                face_lettres = sorted(set(l["Face"] for l in lignes))
                cols_client  = ["Code", "Nom", "Ville", "Quartier", "Adresse"] + [f"Face {l}" for l in face_lettres]
                rows_client  = list(panneaux_client.values())
                df_client    = pd.DataFrame(rows_client, columns=cols_client)

                df_client.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
                ws_c = writer.sheets[sheet_name]

                # Titre
                ws_c["A1"] = f"Panneaux de {client_nom} — {date.today():%d/%m/%Y}"
                ws_c["A1"].font = titre_font

                # En-têtes
                for cell in ws_c[3]:
                    cell.fill      = header_fill
                    cell.font      = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = _thin_border("1E293B")
                ws_c.row_dimensions[3].height = 22

                face_col_indices_client = [
                    i + 1 for i, col in enumerate(cols_client) if col.startswith("Face ")
                ]

                for row in ws_c.iter_rows(min_row=4, max_row=ws_c.max_row):
                    for cell in row:
                        cell.font      = font_normal
                        cell.fill      = fill_normal
                        cell.border    = border_normal
                        cell.alignment = Alignment(vertical="center")

                    for idx in face_col_indices_client:
                        cell = row[idx - 1]
                        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                        val = str(cell.value) if cell.value else ""
                        if "Occupée" in val:
                            cell.fill = fill_occupe
                            cell.font = font_occupe
                        elif "Réservée" in val:
                            cell.fill = fill_reserve
                            cell.font = font_reserve

                    ws_c.row_dimensions[row[0].row].height = 42

                col_widths_client = {
                    "Code": 10, "Nom": 40, "Ville": 15,
                    "Quartier": 15, "Adresse": 35,
                }
                for i, col_name in enumerate(cols_client, start=1):
                    ws_c.column_dimensions[get_column_letter(i)].width = col_widths_client.get(col_name, 28)

                ws_c.freeze_panes = "F4"

        buffer.seek(0)

        parts = ["etat_panneaux"]
        if filters.get("ville"):      parts.append(filters["ville"].replace(" ", "_"))
        if filters.get("quartier"):   parts.append(filters["quartier"].replace(" ", "_"))
        if filters.get("occupation"): parts.append(filters["occupation"])
        parts.append(f"{datetime.now():%Y%m%d}")
        filename = "_".join(parts) + ".xlsx"

        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response















