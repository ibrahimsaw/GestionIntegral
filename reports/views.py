import datetime
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied  # CORRECTION : import manquant
from accounts.decorators import *
from campaigns.models import Campagne, Client
from django.http import HttpResponse
from calendar import monthrange
from django.template.loader import render_to_string
from dateutil.relativedelta import relativedelta
from weasyprint import HTML
import io
from django.views import View

import pandas as pd

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string



class ReportsIndexView(LoginRequiredMixin, View):
    def get(self, request):
        campagnes = Campagne.objects.select_related('client').order_by('-date_debut')
        if request.user.is_client_role and request.user.client_profile:
            campagnes = campagnes.filter(client=request.user.client_profile)
        return render(request, 'reports/index.html', {'campagnes': campagnes})


def _check_campagne_permission(request, campagne):
    """Helper : lève PermissionDenied si l'utilisateur n'a pas accès à la campagne."""
    if request.user.is_client_role:
        if not (request.user.client_profile and campagne.client == request.user.client_profile):
            raise PermissionDenied
    elif not request.user.is_staff:
        raise PermissionDenied


"""
Export PDF du détail d'une campagne — version WeasyPrint.
Gère les campagnes écran ET panneau.
"""



DEFAULT_PLAGE = "06:00-22:00"

MOIS_FR = {
    1: "JANVIER",   2: "FÉVRIER",  3: "MARS",      4: "AVRIL",
    5: "MAI",       6: "JUIN",     7: "JUILLET",   8: "AOÛT",
    9: "SEPTEMBRE", 10: "OCTOBRE", 11: "NOVEMBRE", 12: "DÉCEMBRE",
}

# ══════════════════════════════════════════════════════════════════
# UTILITAIRES
# ══════════════════════════════════════════════════════════════════

def _format_freq(campagne):
    freq = getattr(campagne, "frequence", None)
    if not freq:
        return "1/2 mn"
    if freq < 60:
        return f"Toutes les {freq} sec"
    minutes = freq // 60
    return f"Toutes les {minutes} min"


def _spots_du_mois(campagne, spots_total, annee, mois):
    """
    Calcule les spots d'une campagne pour un mois donné.
    Proratise selon le nombre de jours de la campagne dans ce mois.
    """
    debut = campagne.date_debut
    fin   = campagne.date_fin

    mois_debut = datetime.date(annee, mois, 1)
    mois_fin   = datetime.date(annee, mois, monthrange(annee, mois)[1])

    inter_debut = max(debut, mois_debut)
    inter_fin   = min(fin,   mois_fin)

    if inter_debut > inter_fin:
        return 0

    jours_dans_mois    = (inter_fin - inter_debut).days + 1
    duree_totale_jours = (fin - debut).days + 1

    if duree_totale_jours == 0:
        return 0

    return round(spots_total * jours_dans_mois / duree_totale_jours)


def _mois_couverts(date_debut, date_fin):
    """Retourne la liste des (annee, mois) couverts entre deux dates."""
    mois    = []
    courant = datetime.date(date_debut.year, date_debut.month, 1)
    fin     = datetime.date(date_fin.year,   date_fin.month,   1)
    while courant <= fin:
        mois.append((courant.year, courant.month))
        courant += relativedelta(months=1)
    return mois



# ══════════════════════════════════════════════════════════════════
# CONSTRUCTION DU CONTEXTE CAMPAGNE
# ══════════════════════════════════════════════════════════════════

def _build_context_campagne(campagne):
    """
    Construit le contexte complet pour le PDF d'une campagne.

    Retourne :
    {
        'campagne'      : <Campagne>,
        'client'        : <Client>,
        'today'         : date,
        'infos'         : { ... },   # infos générales
        'supports'      : [ ... ],   # détail supports/écrans/faces
        'spots_par_mois': [ ... ],   # spots proratisés par mois (écran seulement)
    }
    """

    enfants = []
    campagnes = [campagne]

    if campagne.est_mere:
        enfants = list(campagne.sous_campagnes.all())
        campagnes = enfants

    # ── Infos générales ──────────────────────────────────────────
    infos = {
        "nom"                   : campagne.nom,
        "reference"             : campagne.reference,
        "date_debut"            : campagne.date_debut,
        "date_fin"              : campagne.date_fin,
        "duree_jours"           : campagne.duree_jours(),
        "statut"                : campagne.get_statut_display(),
        "type_support"          : "Campagne mère" if campagne.est_mere else campagne.get_type_support_display(),
        "effective_type_support": None,
        "notes"                 : campagne.notes,
        "contrat"               : campagne.contrat,
        "client"                : campagne.client,
        "campagne_mere"         : campagne.est_mere,
        "child_campaigns"       : enfants if campagne.est_mere else None,
        "child_count"           : len(enfants) if campagne.est_mere else 0,
        # Champs écran
        "frequence"      : campagne.frequence,
        "freq_display"   : _format_freq(campagne),
        "duree_passage"  : campagne.duree_passage,
        "tranches"       : campagne.tranches_horaires or "—",
    }

    # ── Lignes associées ─────────────────────────────────────────
    lignes = []
    if campagne.est_mere:
        for enfant in enfants:
            lignes.extend(list(enfant.lignes.all()))
    else:
        lignes = list(campagne.lignes.all())

    effective_support_types = {
        ligne.support.type_support
        for ligne in lignes
        if ligne.support and ligne.support.type_support
    }

    if campagne.type_support == "ecran":
        infos["effective_type_support"] = "ecran"
    elif campagne.type_support == "panneau":
        infos["effective_type_support"] = "panneau"
    elif "ecran" in effective_support_types:
        infos["effective_type_support"] = "ecran"
    elif "panneau" in effective_support_types:
        infos["effective_type_support"] = "panneau"
    else:
        infos["effective_type_support"] = ""

    # ── Détail supports ──────────────────────────────────────────
    supports = []

    if infos["effective_type_support"] == "ecran":
        support_dict = {}
        for ligne in lignes:
            if not ligne.support or ligne.support.type_support != "ecran":
                continue
            support = ligne.support
            support_dict.setdefault(support.pk, {
                "code"   : support.code,
                "nom"    : support.nom,
                "ville"  : support.ville,
                "quartier": support.quartier,
                "adresse": support.adresse,
                "type"   : "Écran",
                "face"   : None,
            })
        supports = list(support_dict.values())

    elif infos["effective_type_support"] == "panneau":
        faces_dict = {}
        for ligne in lignes:
            if not ligne.support or ligne.support.type_support != "panneau":
                continue
            key = ligne.support.pk
            if key not in faces_dict:
                faces_dict[key] = {
                    "code"        : ligne.support.code,
                    "nom"         : ligne.support.nom,
                    "ville"       : ligne.support.ville,
                    "quartier"    : ligne.support.quartier,
                    "adresse"     : ligne.support.adresse,
                    "type"        : "Panneau",
                    "face_labels" : [],
                }
            if ligne.face and ligne.face.label not in faces_dict[key]["face_labels"]:
                faces_dict[key]["face_labels"].append(ligne.face.label)

        for v in faces_dict.values():
            supports.append({
                "code" : v["code"],
                "nom"  : v["nom"],
                "ville" : v["ville"],
                "quartier" : v["quartier"],
                "adresse" : v["adresse"],
                "type" : v["type"],
                "face" : " & ".join(v["face_labels"]) if v["face_labels"] else "—",
            })

    supports_by_child = []
    if campagne.est_mere:
        for enfant in enfants:
            child_supports = []
            if infos["effective_type_support"] == "ecran":
                support_dict = {}
                for ligne in enfant.lignes.all():
                    if not ligne.support or ligne.support.type_support != "ecran":
                        continue
                    support = ligne.support
                    support_dict.setdefault(support.pk, {
                        "code"   : support.code,
                        "nom"    : support.nom,
                        "ville"  : support.ville,
                        "quartier": support.quartier,
                        "adresse": support.adresse,
                        "type"   : "Écran",
                        "face"   : None,
                    })
                child_supports = list(support_dict.values())
            elif infos["effective_type_support"] == "panneau":
                faces_dict = {}
                for ligne in enfant.lignes.all():
                    if not ligne.support or ligne.support.type_support != "panneau":
                        continue
                    key = ligne.support.pk
                    if key not in faces_dict:
                        faces_dict[key] = {
                            "code"        : ligne.support.code,
                            "nom"         : ligne.support.nom,
                            "ville"       : ligne.support.ville,
                            "quartier"    : ligne.support.quartier,
                            "adresse"     : ligne.support.adresse,
                            "type"        : "Panneau",
                            "face_labels" : [],
                        }
                    if ligne.face and ligne.face.label not in faces_dict[key]["face_labels"]:
                        faces_dict[key]["face_labels"].append(ligne.face.label)

                child_supports = [
                    {
                        "code" : v["code"],
                        "nom"  : v["nom"],
                        "ville" : v["ville"],
                        "quartier" : v["quartier"],
                        "adresse" : v["adresse"],
                        "type" : v["type"],
                        "face" : " & ".join(v["face_labels"]) if v["face_labels"] else "—",
                    }
                    for v in faces_dict.values()
                ]

            supports_by_child.append({
                "campagne" : enfant,
                "supports" : child_supports,
            })

    # ── Spots par mois (écran uniquement) ────────────────────────
    spots_par_mois = []

    if infos["effective_type_support"] == "ecran":
        spots_by_month = {}
        total_spots = 0
        for campagne_item in campagnes:
            spots_total = campagne_item.calculer_nombre_spots()
            total_spots += spots_total
            for (annee, mois) in _mois_couverts(campagne_item.date_debut, campagne_item.date_fin):
                spots_mois = _spots_du_mois(campagne_item, spots_total, annee, mois)
                if spots_mois == 0:
                    continue
                spots_by_month[(annee, mois)] = spots_by_month.get((annee, mois), 0) + spots_mois

        nb_ecrans = len({
            ligne.support.pk
            for ligne in lignes
            if ligne.support and ligne.support.type_support == "ecran"
        })

        total_general = sum(spots_by_month.values())
        for (annee, mois), spots_mois in sorted(spots_by_month.items()):
            spots_par_mois.append({
                "label"      : f"{MOIS_FR[mois]} {annee}",
                "spots"      : spots_mois,
                "spots_ecran": spots_mois // nb_ecrans if nb_ecrans else 0,
                "nb_ecrans"  : nb_ecrans,
            })

        if spots_par_mois:
            spots_par_mois.append({
                "label"      : "TOTAL",
                "spots"      : total_general,
                "spots_ecran": total_general // nb_ecrans if nb_ecrans else 0,
                "nb_ecrans"  : nb_ecrans,
                "is_total"   : True,
            })

    infos["support_count"] = len(supports)
    infos["total_spots"] = sum(c.calculer_nombre_spots() for c in campagnes)

    return {
        "campagne"      : campagne,
        "client"        : campagne.client,
        "today"         : datetime.date.today(),
        "infos"         : infos,
        "supports"      : supports,
        "supports_by_child": supports_by_child,
        "spots_par_mois": spots_par_mois,
    }


# ══════════════════════════════════════════════════════════════════
# VUES
# ══════════════════════════════════════════════════════════════════

class ExportCampagnePdfView(ClientStaffRequiredMixin, View):
    def get(self, request, pk):
        """Télécharge le détail d'une campagne en PDF."""
        campagne = get_object_or_404(
            Campagne.objects.select_related("client", "contrat")
                            .prefetch_related(
                                "lignes__support",
                                "lignes__face",
                                "sous_campagnes__lignes__support",
                                "sous_campagnes__lignes__face",
                            ),
            pk=pk,
        )

        html_string = render_to_string(
            "reports/campagne_pdf.html",
            _build_context_campagne(campagne),
            request=request,
        )
        pdf = HTML(
            string=html_string,
            base_url=request.build_absolute_uri(),
        ).write_pdf()

        filename = (
            f"campagne_{campagne.reference}"
            f"_{datetime.datetime.now():%Y%m%d}.pdf"
        )
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

class PreviewCampagnePdfView(ClientStaffRequiredMixin, View):
    def get(self, request, pk):
        """Prévisualise le détail d'une campagne dans le navigateur."""
        campagne = get_object_or_404(
            Campagne.objects.select_related("client", "contrat")
                            .prefetch_related(
                                "lignes__support",
                                "lignes__face",
                                "sous_campagnes__lignes__support",
                                "sous_campagnes__lignes__face",
                            ),
            pk=pk,
        )
        return render(
            request,
            "reports/apercu_campagne.html",
            _build_context_campagne(campagne),
        )

"""
Export PDF du planning de diffusion — version WeasyPrint.

La logique de calcul reste identique, mais tout le rendu
est délégué au template planning_pdf.html.
"""



# ══════════════════════════════════════════════════════════════════
# CONSTRUCTION DES BLOCS ÉCRANS
# ══════════════════════════════════════════════════════════════════

def _build_blocs_ecran(client):
    """
    Structure retournée :

    blocs = [
        {
            'contrat'     : <Contrat> ou None,
            'label_total' : str,
            'total_spots' : int,
            'reste_final' : int,
            'mois'        : [
                {
                    'label'       : "JANVIER 2025",
                    'total_spots' : int,
                    'reste_fin'   : int,
                    'lignes'      : [
                        {
                            'num'          : int,
                            'campagne'     : <Campagne>,
                            'freq_display' : str,
                            'nb_ecrans'    : int,
                            'spots'        : int,
                            'spots_total'  : int,
                            'reste'        : int,
                            'reste_display': str,
                            'ecrans_str'   : str,
                        },
                    ],
                },
            ],
        },
    ]
    """
    blocs   = []
    counter = 1

    # ── Contrats ──────────────────────────────────────────────────
    contrats = (
        client.contrats
        .order_by("date_debut")
        .prefetch_related("campagnes__lignes__support")
    )

    for contrat in contrats:
        campagnes = list(
            contrat.campagnes
            .filter(lignes__support__type_support="ecran")
            .distinct()
            .order_by("date_debut")
        )
        if not campagnes:
            continue

        spots_par_campagne = {
            c.pk: c.calculer_nombre_spots() for c in campagnes
        }

        tous_les_mois = sorted({
            ym
            for c in campagnes
            for ym in _mois_couverts(c.date_debut, c.date_fin)
        })

        reste               = contrat.nb_spots
        total_spots_contrat = 0
        mois_blocs          = []

        for (annee, mois) in tous_les_mois:
            label_mois       = f"{MOIS_FR[mois]} {annee}"
            lignes_mois      = []
            total_spots_mois = 0

            for campagne in campagnes:
                spots_mois = _spots_du_mois(
                    campagne, spots_par_campagne[campagne.pk], annee, mois
                )
                if spots_mois == 0:
                    continue

                # Utilise les données préchargées — évite le N+1
                lignes_ecrans = [
                    l for l in campagne.lignes.all()
                    if l.support.type_support == "ecran"
                ]
                ecrans_str = " • ".join(l.support.nom for l in lignes_ecrans)
                nb_ecrans  = len(lignes_ecrans)

                total_spots_mois    += spots_mois
                total_spots_contrat += spots_mois
                reste               -= spots_mois

                lignes_mois.append({
                    "num"          : counter,
                    "campagne"     : campagne,
                    "freq_display" : _format_freq(campagne),
                    "nb_ecrans"    : nb_ecrans,
                    "spots"        : spots_mois,
                    "spots_total"  : spots_par_campagne[campagne.pk],
                    "reste"        : reste,
                    "reste_display": f"{reste:,}",
                    "ecrans_str"   : lignes_ecrans,
                })
                counter += 1

            if lignes_mois:
                mois_blocs.append({
                    "label"       : label_mois,
                    "total_spots" : total_spots_mois,
                    "reste_fin"   : reste,
                    "lignes"      : lignes_mois,
                })

        blocs.append({
            "contrat"    : contrat,
            "label_total": (contrat.nom or contrat.get_type_contrat_display()).upper(),
            "total_spots": total_spots_contrat,
            "reste_final": reste,
            "mois"       : mois_blocs,
        })

    # ── Campagnes sans contrat ────────────────────────────────────
    campagnes_orphelines = list(
        Campagne.objects
        .filter(
            client=client,
            contrat__isnull=True,
            lignes__support__type_support="ecran",
        )
        .distinct()
        .order_by("date_debut")
        .prefetch_related("lignes__support")
    )

    if campagnes_orphelines:
        spots_par_campagne = {
            c.pk: c.calculer_nombre_spots() for c in campagnes_orphelines
        }

        tous_les_mois = sorted({
            ym
            for c in campagnes_orphelines
            for ym in _mois_couverts(c.date_debut, c.date_fin)
        })

        total_spots_section = 0
        mois_blocs          = []

        for (annee, mois) in tous_les_mois:
            label_mois       = f"{MOIS_FR[mois]} {annee}"
            lignes_mois      = []
            total_spots_mois = 0

            for campagne in campagnes_orphelines:
                spots_mois = _spots_du_mois(
                    campagne, spots_par_campagne[campagne.pk], annee, mois
                )
                if spots_mois == 0:
                    continue

                # Utilise les données préchargées — évite le N+1
                lignes_ecrans = [
                    l for l in campagne.lignes.all()
                    if l.support.type_support == "ecran"
                ]
                ecrans_str = " • ".join(l.support.nom for l in lignes_ecrans)
                nb_ecrans  = len(lignes_ecrans)

                total_spots_mois    += spots_mois
                total_spots_section += spots_mois

                lignes_mois.append({
                    "num"          : counter,
                    "campagne"     : campagne,
                    "freq_display" : _format_freq(campagne),
                    "nb_ecrans"    : nb_ecrans,
                    "spots"        : spots_mois,
                    "spots_total"  : spots_par_campagne[campagne.pk],
                    "reste"        : None,
                    "reste_display": "",
                    "ecrans_str"   : lignes_ecrans
                })
                counter += 1

            if lignes_mois:
                mois_blocs.append({
                    "label"       : label_mois,
                    "total_spots" : total_spots_mois,
                    "reste_fin"   : None,
                    "lignes"      : lignes_mois,
                })

        blocs.append({
            "contrat"    : None,
            "label_total": "SANS CONTRAT",
            "total_spots": total_spots_section,
            "reste_final": None,
            "mois"       : mois_blocs,
        })

    return blocs


# ══════════════════════════════════════════════════════════════════
# CONSTRUCTION DES BLOCS PANNEAUX
# ══════════════════════════════════════════════════════════════════

def _build_blocs_panneaux(client):
    """
    Structure des campagnes panneau d'un client.
    Les faces sont regroupées par support (panneau).
    """
    campagnes = list(
        Campagne.objects
        .filter(client=client, type_support="panneau")
        .distinct()
        .order_by("date_debut")
        .prefetch_related("lignes__support", "lignes__face")
    )

    if not campagnes:
        return []

    lignes  = []
    counter = 1

    for campagne in campagnes:
        lignes_panneau = [
            l for l in campagne.lignes.all()
            if l.support.type_support == "panneau"
        ]

        # Regrouper les faces par support (évite les doublons de nom)
        faces_dict = {}
        for l in lignes_panneau:
            key = l.support.pk
            if key not in faces_dict:
                faces_dict[key] = {
                    "support_code": l.support.code,
                    "support_nom" : l.support.nom,
                    "face_labels" : [],
                }
            if l.face:
                faces_dict[key]["face_labels"].append(l.face.label)

        faces = [
            {
                "support_code": v["support_code"],
                "support_nom" : v["support_nom"],
                "face_label"  : " & ".join(v["face_labels"]) if v["face_labels"] else "-",
            }
            for v in faces_dict.values()
        ]

        nb_faces = sum(len(v["face_labels"]) for v in faces_dict.values())

        lignes.append({
            "num"        : counter,
            "campagne"   : campagne,
            "reference"  : campagne.reference,
            "date_debut" : campagne.date_debut,
            "date_fin"   : campagne.date_fin,
            "duree_jours": campagne.duree_jours(),
            "statut"     : campagne.get_statut_display(),
            "notes"      : campagne.notes,
            "faces"      : faces,
            "nb_faces"   : nb_faces,
        })
        counter += 1

    return [{
        "label"      : "PANNEAUX",
        "total_faces": sum(l["nb_faces"] for l in lignes),
        "lignes"     : lignes,
    }]


# ══════════════════════════════════════════════════════════════════
# VUES
# ══════════════════════════════════════════════════════════════════

def _build_context(client):
    """Contexte partagé entre export PDF et preview."""
    blocs_ecran    = _build_blocs_ecran(client)
    blocs_panneaux = _build_blocs_panneaux(client)

    # ── KPI Écrans ──
    total_campagnes_ecran = sum(
        len(mois["lignes"])
        for bloc in blocs_ecran
        for mois in bloc["mois"]
    )
    total_spots_global = sum(
        bloc["total_spots"] for bloc in blocs_ecran
    )
    reste_spots_global = sum(
        bloc["reste_final"]
        for bloc in blocs_ecran
        if bloc.get("reste_final") is not None
    )

    # ── KPI Panneaux ──
    total_campagnes_panneaux = sum(
        len(bloc["lignes"]) for bloc in blocs_panneaux
    )
    total_faces_global = sum(
        bloc["total_faces"] for bloc in blocs_panneaux
    )

    return {
        "client"                  : client,
        "blocs_ecran"             : blocs_ecran,
        "blocs_panneaux"          : blocs_panneaux,
        "today"                   : datetime.date.today(),
        # KPI conclus
        "total_campagnes_ecran"   : total_campagnes_ecran,
        "total_spots_global"      : total_spots_global,
        "reste_spots_global"      : reste_spots_global,
        "total_campagnes_panneaux": total_campagnes_panneaux,
        "total_faces_global"      : total_faces_global,
        # "DESIGN"                : DESIGN,
    }


class ExportClientPdfView(ClientStaffRequiredMixin, View):
    def get(self, request, pk):
        """Télécharge le planning de diffusion d'un client en PDF."""
        client = get_object_or_404(Client, pk=pk)

        html_string = render_to_string(
            "reports/client_pdf.html",
            _build_context(client),
            request=request,
        )
        pdf = HTML(
            string=html_string,
            base_url=request.build_absolute_uri(),
        ).write_pdf()

        filename = (
            f"rapport_diffusion_{client.nom.replace(' ', '_')}"
            f"_{datetime.datetime.now():%Y%m%d}.pdf"
        )
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class PreviewClientPdfView(ClientStaffRequiredMixin, View):
    def get(self, request, pk):
        """Prévisualise le planning de diffusion dans le navigateur."""
        client = get_object_or_404(Client, pk=pk)
        return render(request, "reports/apercu_client.html", _build_context(client))


class ExportClientExcelView(ClientStaffRequiredMixin, View):
    def get(self, request, pk):
        client = get_object_or_404(Client, pk=pk)

        html_string = render_to_string(
            "reports/client_pdf.html",
            _build_context(client),
            request=request,
        )

        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            tables = pd.read_html(io.StringIO(html_string))

            sheet_names = ["Écrans", "Panneaux"]

            for i, df in enumerate(tables):
                sheet = sheet_names[i] if i < len(sheet_names) else f"Tableau_{i+1}"

                df = df[df.apply(lambda row: row.nunique() > 1, axis=1)]

                df.to_excel(writer, index=False, sheet_name=sheet, startrow=2)

                ws = writer.sheets[sheet]

                ws["A1"] = f"Rapport de diffusion – {client.nom}"
                from openpyxl.styles import Font, PatternFill, Alignment
                ws["A1"].font = Font(bold=True, size=13, name="Arial")

                header_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
                header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                for cell in ws[3]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                from openpyxl.utils import get_column_letter
                for col in ws.columns:
                    max_len = max(
                        (len(str(c.value)) for c in col if c.value is not None),
                        default=10,
                    )
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

                ws.freeze_panes = "A4"

        buffer.seek(0)

        filename = (
            f"rapport_diffusion_{client.nom.replace(' ', '_')}"
            f"_{datetime.datetime.now():%Y%m%d}.xlsx"
        )
        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    
    
# je veux construire un rapport pour tous les panneaux de leur etat actuel en fonction de leur statut et de leur ville, quartier, adresse, etc. et leur nombre de faces ocupées et libres, et les campagnes associées à chaque panneau. Je veux aussi pouvoir filtrer par ville, quartier, statut, etc.






